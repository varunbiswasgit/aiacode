import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os
import sys
import re
from pathlib import Path


def sanitize_filename(filename):
    """
    Remove or replace characters invalid in Windows/Unix file names.
    Returns a safe string suitable for use as a filename stem.
    """
    sanitized = re.sub(r'[<>:"/\\|?*]', '_', str(filename))
    sanitized = sanitized.strip('. ')

    # Full Windows reserved device names
    reserved = {
        'con', 'prn', 'aux', 'nul',
        'com1', 'com2', 'com3', 'com4', 'com5', 'com6', 'com7', 'com8', 'com9',
        'lpt1', 'lpt2', 'lpt3', 'lpt4', 'lpt5', 'lpt6', 'lpt7', 'lpt8', 'lpt9'
    }
    if not sanitized or sanitized.lower() in reserved:
        sanitized = 'Unknown_Manager'

    return sanitized[:200]  # Guard against excessively long names


def autofit_columns(filepath):
    """
    Open an existing .xlsx file and set each column width to fit its content.
    Width is clamped between 8 and 50 characters.
    """
    wb = load_workbook(filepath)
    ws = wb.active
    for idx, col in enumerate(ws.columns, 1):
        max_len = max(
            (len(str(cell.value)) for cell in col if cell.value is not None),
            default=0
        )
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 8), 50)
    wb.save(filepath)
    wb.close()


def split_excel_by_manager(path, manager_column='Manager', output_dir='manager_reports'):
    """
    Split an Excel file into per-manager workbooks.

    Args:
        path (str):             Path to the input Excel file.
        manager_column (str):   Column header that contains manager names.
        output_dir (str):       Directory where output files are saved.

    Returns:
        bool: True if at least one report was written successfully.
    """
    path = Path(path)

    # --- Input validation ---
    if not path.exists():
        print(f"Error: '{path}' does not exist.")
        return False

    if path.suffix.lower() not in ('.xlsx', '.xls', '.xlsm'):
        print(f"Error: Unsupported file type '{path.suffix}'. Accepted: .xlsx, .xls, .xlsm")
        return False

    print(f"Reading: {path}")
    try:
        df = pd.read_excel(path)
    except PermissionError:
        print(f"Error: '{path}' is locked. Close it in Excel and retry.")
        return False
    except Exception as exc:
        print(f"Error reading file: {exc}")
        return False

    if df.empty:
        print("Error: The file contains no data.")
        return False

    if manager_column not in df.columns:
        print(f"Error: Column '{manager_column}' not found.")
        print(f"Available columns: {', '.join(df.columns.tolist())}")
        return False

    valid_rows = df[manager_column].dropna()
    if valid_rows.empty:
        print(f"Error: Column '{manager_column}' has no non-null values.")
        return False

    unique_managers = df[manager_column].dropna().unique()
    print(f"Found {len(unique_managers)} unique manager(s).")

    # --- Output directory ---
    output_path = Path(output_dir)
    try:
        output_path.mkdir(parents=True, exist_ok=True)
    except PermissionError:
        print(f"Error: Cannot create output directory '{output_dir}'.")
        return False

    # --- Per-manager export ---
    success_count = 0
    groups = list(df.groupby(manager_column))

    for manager, group in groups:
        if pd.isna(manager) or str(manager).strip() == '':
            print("Skipping blank manager entry.")
            continue

        safe_name = sanitize_filename(manager)
        out_file = output_path / f"{safe_name}_report.xlsx"
        print(f"  Writing: {manager!r} -> {out_file}")

        try:
            group.to_excel(out_file, index=False)
        except PermissionError:
            print(f"  Error: '{out_file}' is open elsewhere. Skipping.")
            continue
        except Exception as exc:
            print(f"  Error writing '{out_file}': {exc}")
            continue

        try:
            autofit_columns(out_file)
        except Exception as exc:
            print(f"  Warning: Column auto-fit failed for '{out_file}': {exc}")
            # File is still valid; count it as a success

        success_count += 1

    print(f"\nDone. {success_count}/{len(groups)} report(s) saved to: {output_path.resolve()}")
    return success_count > 0


def main():
    """CLI entry point."""
    if len(sys.argv) < 2:
        print("Split Excel by Manager")
        print("=" * 40)
        print("Usage:  python split_excel_by_manager.py <file> [column] [output_dir]")
        print("\nArguments:")
        print("  file        Path to the input Excel file (.xlsx, .xls, .xlsm)")
        print("  column      Manager column header (default: 'Manager')")
        print("  output_dir  Output folder name      (default: 'manager_reports')")
        print("\nExamples:")
        print("  python split_excel_by_manager.py data.xlsx")
        print("  python split_excel_by_manager.py staff.xlsx Supervisor")
        print("  python split_excel_by_manager.py staff.xlsx Supervisor ./output")
        sys.exit(1)

    input_file  = sys.argv[1]
    manager_col = sys.argv[2] if len(sys.argv) > 2 else 'Manager'
    out_dir     = sys.argv[3] if len(sys.argv) > 3 else 'manager_reports'

    print("Split Excel by Manager")
    print("=" * 40)
    print(f"Input file    : {input_file}")
    print(f"Manager column: {manager_col}")
    print(f"Output folder : {out_dir}")
    print("-" * 40)

    if split_excel_by_manager(input_file, manager_col, out_dir):
        print("\n✅ Completed successfully.")
        sys.exit(0)
    else:
        print("\n❌ Operation failed.")
        sys.exit(1)


if __name__ == '__main__':
    main()
